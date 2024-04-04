import csv
import os
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd


def get_downloads_folder():
    home = str(Path.home())
    return os.path.join(home, 'Downloads')


def read_df_from_downloads(filename):
    downloads_folder = get_downloads_folder()
    jobs_csv_path = os.path.join(downloads_folder, filename)
    all_jobs = pd.read_csv(jobs_csv_path)
    return all_jobs


def save_df_to_downloads(df, filename):
    today = datetime.today().strftime('%Y-%m-%d-%H-%M-%S')
    filename = f"{filename}_{today}.csv"
    downloads_folder = get_downloads_folder()
    jobs_csv_path = os.path.join(downloads_folder, filename)

    print(f"Saving jobs to {jobs_csv_path}")
    df.to_csv(jobs_csv_path, quoting=csv.QUOTE_NONNUMERIC, escapechar="\\", index=False)


def save_df_to_downloads_xlsx(df, filename):
    today = datetime.today().strftime('%Y-%m-%d-%H-%M-%S')
    filename = f"{filename}_{today}.xlsx"
    downloads_folder = get_downloads_folder()  # Make sure this returns the correct path
    jobs_xlsx_path = os.path.join(downloads_folder, filename)

    print(f"Saving jobs to {jobs_xlsx_path}")
    with pd.ExcelWriter(jobs_xlsx_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

        workbook = writer.book
        worksheet = workbook[workbook.sheetnames[0]]

        # Set column width
        approx_char_width = 6 / 0.14  # Approximation, depends on font and zoom
        for col in worksheet.columns:
            worksheet.column_dimensions[col[0].column_letter].width = approx_char_width

        row_height = 3 * 72  # 2 inches in points
        for row in worksheet.iter_rows():
            worksheet.row_dimensions[row[0].row].height = row_height

            # Apply text wrap to specific cells
            for cell in row:
                if worksheet.cell(row=1, column=cell.column).value in ["short_summary", "hard_requirements",
                                                                       "description"]:
                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    print(f"File saved to {jobs_xlsx_path}")
